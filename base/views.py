
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from .models import Candidate, Score, ProcedureStation, Activity, Session,VivaScore, FinalGrade
from django.db.models import Prefetch, Sum
import json
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse,HttpResponse
from django.urls import reverse
from django.contrib import messages
from django.forms import modelform_factory
from django.db import transaction
from django.template.loader import get_template
from xhtml2pdf import pisa
import io
from django.views.decorators.http import require_POST
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required

# Add this import at the top with your other imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from .models import GlobalSettings
from django.db.models import Q

def loginPage(request):
    page = 'login'

    if request.user.is_authenticated:
        return redirect('dashboard')
      
    if request.method == 'POST':
        username = request.POST.get('username').lower()
        password = request.POST.get('password') 

        try:
            user = User.objects.get(username= username)
        except:
            messages.error(request, 'User does not exist')
        
        user = authenticate(request, username=username, password=password)


        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Userername OR password does not exist')

        

    context = {'page': page}
    return render(request, 'base/login_register.html', context)

def logoutUser(request):
    logout(request)
    return redirect('dashboard')

def registerPage(request):
    form = UserCreationForm()
    page = 'register'
   
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = user.username.lower()
            user.save()
            login(request,user)
            return redirect('dashboard')
        else:
            messages.error(request, 'An error occurred during registration')

    return render(request, 'base/login_register.html', {'form': form})

@login_required(login_url='login')
def dashboard(request):
    selected_matric = request.GET.get('matric_number')
    selected_station_id = request.GET.get('station_id')

    selected_candidate = None
    selected_station = None
    activities_data = []

 # Get global settings
    global_settings = GlobalSettings.objects.first()
    
    candidates = Candidate.objects.all()
    stations = ProcedureStation.objects.all()

    if global_settings:
        if global_settings.active_session:
            candidates = candidates.filter(session=global_settings.active_session)
            stations = stations.filter(session=global_settings.active_session)
            
        if global_settings.active_level:
            candidates = candidates.filter(level=global_settings.active_level)
            stations = stations.filter(
                Q(level=global_settings.active_level) | 
                Q(level__isnull=True)
            )

    if selected_station_id:
        selected_station = get_object_or_404(ProcedureStation, id=selected_station_id)

        # If candidate is selected, show scores
        if selected_matric:
            selected_candidate = get_object_or_404(Candidate, matric_number=selected_matric)

            if not selected_station and selected_candidate.procedure_stations.exists():
                selected_station = selected_candidate.procedure_stations.first()

            if selected_station:
                for activity in selected_station.activities.all():
                    score_obj = Score.objects.filter(candidate=selected_candidate, activity=activity).first()
                    options = activity.get_score_options()
                    activities_data.append({
                        'activity_id': activity.id,
                        'description': activity.description,
                        'max_score': activity.max_score,
                        'score': score_obj.score if score_obj else None,
                        'options': options
                    })
        else:
            # Candidate not selected: show activities without scores
            for activity in selected_station.activities.all():
                activities_data.append({
                    'activity_id': activity.id,
                    'description': activity.description,
                    'max_score': activity.max_score,
                    'score': None,
                    'options': activity.get_score_options()
                })
    

    

    context = {
        'candidates': candidates,
        'selected_candidate': selected_candidate,
        'stations': stations,
        'selected_station': selected_station,
        'activities': activities_data,
        'global_settings': global_settings,
    }

    return render(request, 'base/home.html', context)




@login_required(login_url='login')
@csrf_exempt
def save_scores(request):
    if request.method == 'POST':
        matric_number = request.POST.get('matric_number')
        station_id = request.POST.get('station_id')

        candidate = get_object_or_404(Candidate, matric_number=matric_number)

        for key, value in request.POST.items():
            if key.startswith('score_'):
                activity_id = key.split('_')[1]
                try:
                    activity = Activity.objects.get(id=activity_id)
                    score_val = float(value)

                    Score.objects.update_or_create(
                        candidate=candidate,
                        activity=activity,
                        defaults={'score': score_val}
                    )
                except (Activity.DoesNotExist, ValueError):
                    continue

        # ✅ Correct redirect with reverse and query string
        
        messages.success(request, f'Scores saved successfully for {matric_number}.')

        dashboard_url = reverse('dashboard')
        return redirect(f'{dashboard_url}?matric_number={matric_number}&station_id={station_id}')

    # Optional fallback
    messages.success(request, f'Scores saved successfully for {matric_number}.')
    return redirect(reverse('dashboard'))



@login_required(login_url='login')
def viva_scoring_view(request):
    global_settings = GlobalSettings.objects.first()
    
    
    candidates = Candidate.objects.all()
    stations = ProcedureStation.objects.all()
    
    if global_settings:
        if global_settings.active_session:
            candidates = candidates.filter(session=global_settings.active_session)
            stations = stations.filter(session=global_settings.active_session)
        if global_settings.active_level:
            candidates = candidates.filter(level=global_settings.active_level)
            stations = stations.filter(
                Q(level=global_settings.active_level) | 
                Q(level__isnull=True)
            )

    if request.method == 'POST':
        with transaction.atomic():
            for candidate in candidates:
                score_value = request.POST.get(f'score_{candidate.id}')
                if score_value:
                    score_value = float(score_value)
                    if 0 <= score_value <= 10:
                        VivaScore.objects.update_or_create(
                            candidate=candidate,
                            defaults={'score': score_value}
                        )
        messages.success(request, 'Scores saved successfully.')
        return redirect('viva_scores')
   
    return render(request, 'base/viva.html', {
        'candidates': candidates,
        'global_settings': global_settings,
        'stations': stations

    })


from django.core.paginator import Paginator

@login_required(login_url='login')
def final_grade_report(request):
    global_settings = GlobalSettings.objects.first()
    
    candidates = Candidate.objects.all()
    stations = ProcedureStation.objects.all()
    
    if global_settings:
        if global_settings.active_session:
            candidates = candidates.filter(session=global_settings.active_session)
            stations = stations.filter(session=global_settings.active_session)
        if global_settings.active_level:
            candidates = candidates.filter(level=global_settings.active_level)
            stations = stations.filter(
                Q(level=global_settings.active_level) | 
                Q(level__isnull=True)
            )

    report_data = []
    for candidate in candidates:
        viva_score = VivaScore.objects.filter(candidate=candidate).first()
        viva = viva_score.score if viva_score else 0

        final_grade, created = FinalGrade.objects.get_or_create(
            candidate=candidate, 
            defaults={'viva': viva}
        )
        if not created:
            final_grade.viva = viva
        final_grade.save()

        station_scores = []
        for station in stations:
            total_score = Score.objects.filter(
                candidate=candidate,
                activity__station=station
            ).aggregate(Sum('score'))['score__sum'] or 0
            station_scores.append((station.name, total_score))

        report_data.append({
            'matric_number': candidate.matric_number,
            'full_name': candidate.full_name,
            'station_scores': station_scores,
            'viva': viva,
            'total': final_grade.total,
            'level': candidate.get_level_display(),
        })

    # Paginate report_data
    paginator = Paginator(report_data, 20)  # Show 20 candidates per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'base/final_grade_report.html', {
        'page_obj': page_obj,
        'stations': stations,
        'global_settings': global_settings
    })

@login_required(login_url='login')
@require_POST
def recalculate_grades(request):
    global_settings = GlobalSettings.objects.first()
    
    candidates = Candidate.objects.all()
    if global_settings:
        if global_settings.active_session:
            candidates = candidates.filter(session=global_settings.active_session)
        if global_settings.active_level:
            candidates = candidates.filter(level=global_settings.active_level)

    for candidate in candidates:
        viva_score = VivaScore.objects.filter(candidate=candidate).first()
        viva = viva_score.score if viva_score else 0

        final_grade, _ = FinalGrade.objects.get_or_create(candidate=candidate)
        final_grade.viva = viva
        final_grade.save()

    messages.success(request, "Grades recalculated successfully.")
    return redirect('final_grade_report')




@login_required(login_url='login')
def download_grade_report_pdf(request):
    global_settings = GlobalSettings.objects.first()
    
    candidates = Candidate.objects.all()
    stations = ProcedureStation.objects.all()
    
    if global_settings:
        if global_settings.active_session:
            candidates = candidates.filter(session=global_settings.active_session)
            stations = stations.filter(session=global_settings.active_session)
        if global_settings.active_level:
            candidates = candidates.filter(level=global_settings.active_level)
            stations = stations.filter(
                Q(level=global_settings.active_level) | 
                Q(level__isnull=True)
            )
    report_data = []
    for candidate in candidates:
        viva_score = VivaScore.objects.filter(candidate=candidate).first()
        viva = viva_score.score if viva_score else 0
        final_grade = FinalGrade.objects.get(candidate=candidate)
        station_scores = []
        for station in stations:
            total_score = Score.objects.filter(
                candidate=candidate,
                activity__station=station
            ).aggregate(Sum('score'))['score__sum'] or 0
            station_scores.append((station.name, total_score))

        report_data.append({
            'matric_number': candidate.matric_number,
            'full_name': candidate.full_name,
            'station_scores': station_scores,
            'viva': viva,
            'total': final_grade.total
        })

    template = get_template('base/final_grade_report_pdf.html')
    html = template.render({'report_data': report_data, 'stations': stations})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="final_grade_report.pdf"'

    pisa_status = pisa.CreatePDF(io.StringIO(html), dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response


def index(request):
   
    return render(request, 'base/index.html')




from django.http import JsonResponse
from django.template.loader import render_to_string

from django.views.decorators.http import require_GET

@login_required(login_url='login')
@require_GET
def get_activities_ajax(request):
    global_settings = GlobalSettings.objects.first()
    
    selected_matric = request.GET.get('matric_number')
    selected_station_id = request.GET.get('station_id')

    selected_candidate = Candidate.objects.filter(matric_number=selected_matric).first()
    selected_station = ProcedureStation.objects.filter(id=selected_station_id).first()

    # Filter based on global settings
    candidates = Candidate.objects.all()
    stations = ProcedureStation.objects.all()
    
    if global_settings:
        if global_settings.active_session:
            candidates = candidates.filter(session=global_settings.active_session)
            stations = stations.filter(session=global_settings.active_session)
        if global_settings.active_level:
            candidates = candidates.filter(level=global_settings.active_level)
            stations = stations.filter(
                Q(level=global_settings.active_level) | 
                Q(level__isnull=True)
            )


    activities_data = []
    if selected_station:
        for activity in selected_station.activities.all():
            score_obj = Score.objects.filter(candidate=selected_candidate, activity=activity).first() if selected_candidate else None
            activities_data.append({
                'activity_id': activity.id,
                'description': activity.description,
                'max_score': activity.max_score,
                'score': score_obj.score if score_obj else None,
                'options': activity.get_score_options()
            })

    context = {
        'candidates': candidates,
        'stations': stations,
        'selected_candidate': selected_candidate,
        'selected_station': selected_station,
        'activities': activities_data,
    }

    activities_html = render_to_string('base/activities.html', context, request=request)
    candidate_info_html = render_to_string('base/candidate_info.html', context, request=request)

    return JsonResponse({
        'activities_html': activities_html,
        'candidate_info_html': candidate_info_html,
    })



# Add this view function (place it near your PDF download view)
@login_required(login_url='login')
def download_grade_report_excel(request):

    global_settings = GlobalSettings.objects.first()
    
    candidates = Candidate.objects.all()
    stations = ProcedureStation.objects.all()
    
    if global_settings:
        if global_settings.active_session:
            candidates = candidates.filter(session=global_settings.active_session)
            stations = stations.filter(session=global_settings.active_session)
        if global_settings.active_level:
            candidates = candidates.filter(level=global_settings.active_level)
            stations = stations.filter(
                Q(level=global_settings.active_level) | 
                Q(level__isnull=True)
            )
    # Create the Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Grade Report"

    # Create header row
    headers = ["Matric Number", "Full Name"] + [station.name for station in stations] + ["Viva", "Total"]
    ws.append(headers)

    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Add data rows
    for candidate in candidates:
        viva_score = VivaScore.objects.filter(candidate=candidate).first()
        viva = viva_score.score if viva_score else 0
        final_grade = FinalGrade.objects.get(candidate=candidate)
        
        row_data = [candidate.matric_number, candidate.full_name]
        
        # Add station scores
        for station in stations:
            total_score = Score.objects.filter(
                candidate=candidate,
                activity__station=station
            ).aggregate(Sum('score'))['score__sum'] or 0
            row_data.append(total_score)
        
        # Add viva and total
        row_data.extend([viva, final_grade.total])
        ws.append(row_data)

    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.column == ws.max_column:  # Total column
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='center')

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="final_grade_report.xlsx"'
    wb.save(response)

    return response

from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect

@login_required
def lockscreen(request):
    # If user submits password on lockscreen
    if request.method == 'POST':
        password = request.POST.get('password')
        user = authenticate(username=request.user.username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.POST.get('next', 'dashboard')
            return redirect(next_url)
        else:
            # Return to lockscreen with error
            return render(request, 'lockscreen.html', {'form': {'errors': True}})
    
    return render(request, 'base/lockscreen.html')

from django.shortcuts import render, redirect
from .forms import GlobalSettingsForm
from .models import GlobalSettings

def settings_view(request):

    settings = GlobalSettings.objects.first() or GlobalSettings()
    
    if request.method == 'POST':
        form = GlobalSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            return redirect('dashboard')  # or your preferred page
    else:
        form = GlobalSettingsForm(instance=settings)
    
    return render(request, 'base/settings.html', {'form': form})

# your_app/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import BatchUploadForm
from .batch_upload import process_station_csv, process_candidate_csv, process_activity_csv,process_combined_csv


@login_required
def batch_upload(request):
    
    global_settings = GlobalSettings.objects.first()
    
    candidates = Candidate.objects.all()
    stations = ProcedureStation.objects.all()
    
    if global_settings:
        if global_settings.active_session:
            candidates = candidates.filter(session=global_settings.active_session)
            stations = stations.filter(session=global_settings.active_session)
        if global_settings.active_level:
            candidates = candidates.filter(level=global_settings.active_level)
            stations = stations.filter(
                Q(level=global_settings.active_level) | 
                Q(level__isnull=True)
            )
    if request.method == 'POST':
        form = BatchUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                session = form.cleaned_data['session']
                level = form.cleaned_data['level']
                upload_type = form.cleaned_data['upload_type']
                
                if form.cleaned_data['upload_type'] == 'candidates':
                    if not level:
                        raise ValueError("Level is required for candidate upload")
                    process_candidate_csv(
                        request.FILES['csv_file'],
                        session,
                        level
                    )
                    
                elif upload_type == 'stations':
                    process_station_csv(request.FILES['csv_file'], session, level)
                    
                elif upload_type == 'activities':
                    process_activity_csv(request.FILES['csv_file'], session)
                    
                elif upload_type == 'combined':
                    process_combined_csv(request.FILES['csv_file'], session, level)
                
                messages.success(request, f"{upload_type.capitalize()} uploaded successfully!")
                return redirect('dashboard')
                
            except Exception as e:
                messages.error(request, f"Upload failed: {str(e)}")
    else:
        form = BatchUploadForm()
    
    return render(request, 'base/batch_upload.html', {'form': form,'stations': stations})